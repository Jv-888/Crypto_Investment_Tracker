from openpyxl import load_workbook, Workbook
import os
import imaplib
import email
import re
from email.utils import parsedate_to_datetime
import pytz
from bs4 import BeautifulSoup

# === SETTINGS FOR EXCEL & EMAIL ===
xlsx_path = "" #set the path of the excel file to be created or modified
headers = ["Asset", "Reference Code", "Date", "Amount Invested", "Amount of Coins Acquired", "Price Per Coin"] # data that we will extract from emails
gmail_user = os.getenv("GMAIL_USER") # or "insert credentials here"
gmail_pass = os.getenv("GMAIL_APP_PASS") # or "insert app password here"
if not gmail_user or not gmail_pass:
    raise ValueError("GMAIL credentials not set in environment variables.")
mailbox = '"[Gmail]/All Mail"'
search_date = "19-May-2025" #DD-MM-YYYY
search_subject = "purchase is now available to trade"
search_sender = "no-reply@info.coinbase.com"

# === FILE SETUP ===
file_exists = os.path.exists(xlsx_path)
wb = load_workbook(xlsx_path) if file_exists else Workbook()
ws = wb.active
if not file_exists:
    ws.append(headers)

# === EMAIL SETUP ===
mail = imaplib.IMAP4_SSL("imap.gmail.com")
mail.login(gmail_user, gmail_pass)
mail.select(mailbox)

# === SEARCH EMAILS ===
status, message_ids = mail.search(
    None,
    f'(FROM "{search_sender}" SUBJECT "{search_subject}" SENTSINCE "{search_date}")'
)

# === FIELD EXTRACTOR FUNCTION TO EXTRACT DATA NEEDED ===
def extract_field(label, soup):
    cell = soup.find("td", string=lambda s: s and label in s)
    if cell and cell.find_next_sibling("td"):
        return cell.find_next_sibling("td").get_text(strip=True)
    return "N/A"

# === EXISTING REFERENCE CODES TO PREVENT DUPLICATES ===
existing_refs = {cell.value for cell in ws['B'][1:]}
added_count = 0
skipped_count = 0

# === PROCESS EMAILS ===
if status == "OK":
    for num in message_ids[0].split():
        _, data = mail.fetch(num, "(RFC822)")
        raw_email = data[0][1]
        msg = email.message_from_bytes(raw_email)

        # Get HTML part
        html_body = ""
        for part in msg.walk():
            if part.get_content_type() == "text/html":
                html_body = part.get_payload(decode=True).decode()
                break
        soup = BeautifulSoup(html_body, "html.parser")

        # Extract data
        ref_span = soup.find("span", style=lambda s: s and "monospace" in s)
        ref_code = ref_span.get_text(strip=True) if ref_span else "N/A"
        amount = extract_field("Amount", soup)
        price = extract_field("Price", soup).replace("@", "").split()[0]
        subtotal = extract_field("Subtotal", soup)

        asset = re.sub(r"\d+", "", amount).replace(",", "").replace(".", "").strip()
        coin_amt = amount.split()[0]

        # Parse time
        utc_date = parsedate_to_datetime(msg["Date"])
        local_date = utc_date.astimezone(pytz.timezone("America/New_York"))
        formatted_date = local_date.strftime("%Y-%m-%d")

        # Skip if duplicate
        if ref_code in existing_refs:
            print(f"⚠️ Duplicate found: {ref_code}. Skipping.")
            skipped_count += 1
            continue

        # Append row
        ws.append([
            asset,
            ref_code.strip(),
            formatted_date,
            subtotal.strip(),
            coin_amt,
            price
        ])
        print(f"✅ Added entry for {asset} ({ref_code}) - {formatted_date}")
        added_count += 1

    # Save workbook
    wb.save(xlsx_path)
    print(f"\n💾 Saved to {os.path.basename(xlsx_path)}")
    print(f"🔢 Total messages found: {len(message_ids[0].split())}")
    print(f"✅ New entries added: {added_count}")
    print(f"⛔ Skipped duplicates: {skipped_count}")

mail.logout()
